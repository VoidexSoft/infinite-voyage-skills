#!/usr/bin/env python3
"""
Data Modeler Template Generator for Infinite Voyage

Generates production-quality game data spreadsheets (.xlsx) for character stats,
item databases, and economy models. These templates serve as the single source of
truth for game designers, balance teams, and engine programmers.

Usage:
    python generate_templates.py
    python generate_templates.py --output-dir ./output

Requirements:
    pip install openpyxl
"""

import argparse
import os
import sys
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font,
        PatternFill,
        Alignment,
        Border,
        Side,
        numbers,
    )
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "Error: openpyxl is not installed.\n"
        "Install it with:\n"
        "    pip install openpyxl\n"
        "\n"
        "Or if using a virtual environment:\n"
        "    python -m pip install openpyxl"
    )
    sys.exit(1)


# ---------------------------------------------------------------------------
# Shared styling constants
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUBHEADER_FILL = PatternFill(
    start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"
)
SUBHEADER_FONT = Font(bold=True, size=11)

BODY_FONT = Font(size=10)
BODY_ALIGNMENT = Alignment(vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Conditional formatting fills
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

TODAY = datetime.now().strftime("%Y-%m-%d")


def apply_header_row(ws, headers, row=1):
    """Apply styled headers to a worksheet row."""
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_num)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_data_row(ws, row_num, values):
    """Write a data row with standard styling."""
    for col_num, value in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = value
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGNMENT
        cell.border = THIN_BORDER


def auto_column_widths(ws, min_width=10, max_width=30):
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ===========================================================================
# 1. Character Master Database
# ===========================================================================


def generate_character_master(output_dir):
    """Generate character-master.xlsx with stat columns and conditional formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Characters"

    headers = [
        "ID",
        "Name",
        "Class",
        "Level",
        "HP",
        "MP",
        "STR",
        "DEX",
        "INT",
        "WIS",
        "CON",
        "CHA",
        "Attack",
        "Defense",
        "Speed",
        "XP_Required",
    ]
    apply_header_row(ws, headers)

    # Example character data -- 5 rows covering archetypes
    characters = [
        [
            "char_001",
            "Kael the Warden",
            "Warrior",
            1,
            120,
            30,
            16,
            12,
            8,
            10,
            15,
            10,
            28,
            22,
            10,
            0,
        ],
        [
            "char_002",
            "Lyra Starweave",
            "Mage",
            1,
            65,
            100,
            6,
            10,
            18,
            16,
            8,
            14,
            12,
            8,
            11,
            0,
        ],
        [
            "char_003",
            "Shade",
            "Rogue",
            1,
            80,
            50,
            10,
            18,
            10,
            12,
            10,
            12,
            22,
            12,
            16,
            0,
        ],
        [
            "char_004",
            "Brother Aldric",
            "Cleric",
            1,
            90,
            80,
            10,
            8,
            14,
            18,
            14,
            16,
            14,
            16,
            9,
            0,
        ],
        [
            "char_005",
            "Fenris",
            "Ranger",
            1,
            85,
            60,
            12,
            16,
            12,
            14,
            12,
            10,
            24,
            14,
            14,
            0,
        ],
    ]

    for row_idx, row_data in enumerate(characters, 2):
        apply_data_row(ws, row_idx, row_data)

    # --- Conditional formatting for balance ranges ---
    # Stat columns G through L (STR, DEX, INT, WIS, CON, CHA) -- columns 7-12
    stat_range = "G2:L100"

    # Green: balanced (10-15)
    ws.conditional_formatting.add(
        stat_range,
        CellIsRule(
            operator="between",
            formula=["10", "15"],
            fill=GREEN_FILL,
        ),
    )
    # Yellow: slightly off (6-9 or 16-18)
    ws.conditional_formatting.add(
        stat_range,
        CellIsRule(
            operator="between",
            formula=["16", "18"],
            fill=YELLOW_FILL,
        ),
    )
    ws.conditional_formatting.add(
        stat_range,
        CellIsRule(
            operator="between",
            formula=["6", "9"],
            fill=YELLOW_FILL,
        ),
    )
    # Red: extreme (< 6 or > 18)
    ws.conditional_formatting.add(
        stat_range,
        CellIsRule(
            operator="lessThan",
            formula=["6"],
            fill=RED_FILL,
        ),
    )
    ws.conditional_formatting.add(
        stat_range,
        CellIsRule(
            operator="greaterThan",
            formula=["18"],
            fill=RED_FILL,
        ),
    )

    # HP conditional formatting (column E)
    hp_range = "E2:E100"
    ws.conditional_formatting.add(
        hp_range,
        CellIsRule(operator="between", formula=["80", "120"], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        hp_range,
        CellIsRule(operator="lessThan", formula=["60"], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        hp_range,
        CellIsRule(operator="greaterThan", formula=["150"], fill=RED_FILL),
    )

    # Freeze top row and auto-fit columns
    ws.freeze_panes = "A2"
    auto_column_widths(ws)

    # Add a notes sheet explaining the formatting
    notes_ws = wb.create_sheet("Formatting Guide")
    notes_ws["A1"] = "Conditional Formatting Legend"
    notes_ws["A1"].font = Font(bold=True, size=14)
    notes_ws["A3"] = "Stats (STR, DEX, INT, WIS, CON, CHA):"
    notes_ws["A3"].font = Font(bold=True)
    notes_ws["A4"] = "Green = Balanced (10-15)"
    notes_ws["B4"].fill = GREEN_FILL
    notes_ws["A5"] = "Yellow = Slightly Off (6-9 or 16-18)"
    notes_ws["B5"].fill = YELLOW_FILL
    notes_ws["A6"] = "Red = Extreme (< 6 or > 18)"
    notes_ws["B6"].fill = RED_FILL
    notes_ws["A8"] = "HP:"
    notes_ws["A8"].font = Font(bold=True)
    notes_ws["A9"] = "Green = Standard range (80-120)"
    notes_ws["B9"].fill = GREEN_FILL
    notes_ws["A10"] = "Red = Too low (< 60) or too high (> 150)"
    notes_ws["B10"].fill = RED_FILL
    notes_ws.column_dimensions["A"].width = 45
    notes_ws.column_dimensions["B"].width = 12

    filepath = os.path.join(output_dir, "character-master.xlsx")
    wb.save(filepath)
    print(f"  Created: {filepath}")


# ===========================================================================
# 2. Item Database
# ===========================================================================


def generate_item_database(output_dir):
    """Generate item-database.xlsx with 10 example items across rarities."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Items"

    headers = [
        "ID",
        "Name",
        "Type",
        "Rarity",
        "Level_Req",
        "Stat_Bonus",
        "Value",
        "Drop_Rate",
        "Source",
    ]
    apply_header_row(ws, headers)

    items = [
        ["item_sword_001", "Iron Sword", "Weapon", "Common", 1, "+3 ATK", 50, "15%", "Bandit Drop"],
        ["item_shield_001", "Wooden Shield", "Armor", "Common", 1, "+2 DEF", 35, "18%", "Goblin Drop"],
        ["item_bow_001", "Hunter's Longbow", "Weapon", "Uncommon", 5, "+6 ATK, +2 DEX", 180, "8%", "Forest Ranger Drop"],
        ["item_ring_001", "Silver Ring of Focus", "Accessory", "Uncommon", 5, "+4 INT", 200, "7%", "Chest (Ruins)"],
        ["item_staff_001", "Starweave Staff", "Weapon", "Rare", 10, "+10 INT, +5 WIS", 750, "3%", "Void Cultist Drop"],
        ["item_armor_001", "Warden's Plate", "Armor", "Rare", 10, "+8 DEF, +4 CON", 900, "2.5%", "Dungeon Boss"],
        ["item_potion_001", "Elixir of Swiftness", "Consumable", "Rare", 8, "+5 Speed (temp)", 120, "5%", "Alchemist Quest"],
        ["item_dagger_001", "Nightfall Dagger", "Weapon", "Epic", 20, "+15 ATK, +8 DEX, +Crit", 3500, "0.8%", "Shadow Lord Drop"],
        ["item_amulet_001", "Amulet of the Cosmos", "Accessory", "Epic", 20, "+10 INT, +10 WIS", 4200, "0.5%", "Ancient Vault"],
        ["item_blade_001", "Eternity Blade", "Weapon", "Legendary", 30, "+25 ATK, +10 STR, +Life Steal", 15000, "0.1%", "Final Boss"],
    ]

    rarity_fills = {
        "Common": PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid"),
        "Uncommon": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Rare": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
        "Epic": PatternFill(start_color="D9C2EC", end_color="D9C2EC", fill_type="solid"),
        "Legendary": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
    }

    for row_idx, row_data in enumerate(items, 2):
        apply_data_row(ws, row_idx, row_data)
        rarity = row_data[3]
        if rarity in rarity_fills:
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_num).fill = rarity_fills[rarity]

    ws.freeze_panes = "A2"
    auto_column_widths(ws)

    # Rarity Legend sheet
    legend_ws = wb.create_sheet("Rarity Legend")
    legend_ws["A1"] = "Rarity Color Legend"
    legend_ws["A1"].font = Font(bold=True, size=14)
    for idx, (rarity, fill) in enumerate(rarity_fills.items(), 3):
        legend_ws.cell(row=idx, column=1, value=rarity)
        legend_ws.cell(row=idx, column=2).fill = fill
        legend_ws.cell(row=idx, column=2, value="Sample")
    legend_ws.column_dimensions["A"].width = 20
    legend_ws.column_dimensions["B"].width = 15

    filepath = os.path.join(output_dir, "item-database.xlsx")
    wb.save(filepath)
    print(f"  Created: {filepath}")


# ===========================================================================
# 3. Economy Model
# ===========================================================================


def generate_economy_model(output_dir):
    """Generate economy-model.xlsx with Faucets, Sinks, Balance, and Summary sheets."""
    wb = Workbook()

    # ---- Sheet 1: Faucets ----
    ws_faucets = wb.active
    ws_faucets.title = "Faucets"
    faucet_headers = ["Source", "Amount_Per_Hour", "Level_Req"]
    apply_header_row(ws_faucets, faucet_headers)

    faucets = [
        ["Enemy Kills (Standard)", 300, 1],
        ["Enemy Kills (Elite)", 600, 10],
        ["Quest Rewards", 1000, 1],
        ["Chest Loot", 150, 1],
        ["Crafting Salvage", 200, 5],
        ["Daily Login Bonus", 100, 1],
        ["Arena Rewards", 500, 15],
        ["Trading Profits", 250, 10],
    ]
    for row_idx, row_data in enumerate(faucets, 2):
        apply_data_row(ws_faucets, row_idx, row_data)

    ws_faucets.freeze_panes = "A2"
    auto_column_widths(ws_faucets)

    # ---- Sheet 2: Sinks ----
    ws_sinks = wb.create_sheet("Sinks")
    sink_headers = ["Sink", "Cost", "Frequency"]
    apply_header_row(ws_sinks, sink_headers)

    sinks = [
        ["Item Purchase (Shop)", 500, "Per item"],
        ["Gear Repair", 100, "Per death"],
        ["Gear Upgrade", 1500, "Per upgrade"],
        ["Fast Travel", 50, "Per use"],
        ["Crafting Cost", 300, "Per craft"],
        ["Skill Respec", 2000, "Per respec"],
        ["Cosmetic Purchase", 5000, "Per cosmetic"],
        ["Guild Contribution", 200, "Per week"],
    ]
    for row_idx, row_data in enumerate(sinks, 2):
        apply_data_row(ws_sinks, row_idx, row_data)

    ws_sinks.freeze_panes = "A2"
    auto_column_widths(ws_sinks)

    # ---- Sheet 3: Balance ----
    ws_balance = wb.create_sheet("Balance")
    balance_headers = ["Level", "Income_Rate", "Expense_Rate", "Net_Flow"]
    apply_header_row(ws_balance, balance_headers)

    balance_data = [
        [1, 450, 200, None],
        [5, 650, 400, None],
        [10, 1100, 800, None],
        [15, 1600, 1350, None],
        [20, 2100, 1900, None],
        [25, 2800, 2650, None],
        [30, 3500, 3400, None],
        [40, 5000, 4800, None],
        [50, 7000, 6800, None],
    ]

    for row_idx, row_data in enumerate(balance_data, 2):
        apply_data_row(ws_balance, row_idx, row_data[:3])
        cell = ws_balance.cell(row=row_idx, column=4)
        cell.value = f"=B{row_idx}-C{row_idx}"
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGNMENT
        cell.border = THIN_BORDER

    ws_balance.conditional_formatting.add(
        "D2:D100",
        CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL),
    )
    ws_balance.conditional_formatting.add(
        "D2:D100",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
    )
    ws_balance.conditional_formatting.add(
        "D2:D100",
        CellIsRule(operator="equal", formula=["0"], fill=YELLOW_FILL),
    )

    ws_balance.freeze_panes = "A2"
    auto_column_widths(ws_balance)

    # ---- Sheet 4: Summary ----
    ws_summary = wb.create_sheet("Summary")
    ws_summary["A1"] = "Economy Model Summary"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A1"].alignment = Alignment(horizontal="left")

    ws_summary["A3"] = "Generated"
    ws_summary["B3"] = TODAY
    ws_summary["A4"] = "Game"
    ws_summary["B4"] = "Infinite Voyage"

    ws_summary["A6"] = "Key Metrics"
    ws_summary["A6"].font = Font(bold=True, size=13)

    metrics = [
        ["Total Faucet Sources", f"=COUNTA(Faucets!A2:A100)"],
        ["Total Sink Types", f"=COUNTA(Sinks!A2:A100)"],
        ["Avg Income (All Levels)", f"=AVERAGE(Balance!B2:B100)"],
        ["Avg Expense (All Levels)", f"=AVERAGE(Balance!C2:C100)"],
        ["Avg Net Flow", f"=AVERAGE(Balance!D2:D100)"],
        ["Max Net Flow", f"=MAX(Balance!D2:D100)"],
        ["Min Net Flow", f"=MIN(Balance!D2:D100)"],
    ]

    for idx, (label, formula) in enumerate(metrics, 7):
        ws_summary.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=idx, column=2, value=formula).font = BODY_FONT

    ws_summary["A16"] = "Design Notes"
    ws_summary["A16"].font = Font(bold=True, size=13)
    ws_summary["A17"] = (
        "Economy should maintain a slight positive net flow at all levels to "
        "keep players progressing. Late-game sinks (cosmetics, guild fees, respec) "
        "absorb surplus gold to prevent hyperinflation."
    )
    ws_summary["A17"].alignment = Alignment(wrap_text=True)

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 25

    filepath = os.path.join(output_dir, "economy-model.xlsx")
    wb.save(filepath)
    print(f"  Created: {filepath}")


# ===========================================================================
# CLI entry point
# ===========================================================================


def main():
    parser = argparse.ArgumentParser(
        description="Generate Infinite Voyage data modeler spreadsheet templates (.xlsx).",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  python generate_templates.py\n"
            "  python generate_templates.py --output-dir ./output\n"
        ),
    )
    parser.add_argument(
        "--output-dir",
        default=os.path.dirname(os.path.abspath(__file__)),
        help="Directory where .xlsx files will be written (default: script directory).",
    )
    args = parser.parse_args()

    output_dir = os.path.abspath(args.output_dir)
    os.makedirs(output_dir, exist_ok=True)

    print(f"Generating data modeler templates in: {output_dir}\n")

    generate_character_master(output_dir)
    generate_item_database(output_dir)
    generate_economy_model(output_dir)

    print("\nDone. All data modeler templates generated successfully.")


if __name__ == "__main__":
    main()
